from openpyxl import Workbook, load_workbook
from Classes import Customer
from Mappings import ship2addy, bill2addy, headings, name, address, cleaningDay, notes, longitude, latitude, GoodAddress, firstFile, firstOutput, cName
from GoogleStuff import getCoordinates

def FirstLoad():
    #Load workbook, get ready to access data
    workbook = load_workbook(filename=firstFile) #didn't set readonly to true becasue will add lat and lng for potties without them already
    sheet = workbook.active

    #access data and store data in a list of Potty classes
    customers = []
    i = 0
    for row in sheet.iter_rows(min_row=2):
        customer = Customer()
        customer.name = row[cName].value
        #grab string of ship2addy and bill2addy and figure out which one to use
            #ship2addy first
            #find first cell that starts with a number
        addy = ''
        seenNum = False
        for i in ship2addy:
            if row[i].value == None: continue #skipping cells that have None value
            if ord(row[i].value[0]) <= ord('9') or ord(row[i].value[0]) >= ord('0'): #this cell starts with a number, this is checked so that all the cases where the customer name is first-second cell get skipped
                seenNum = True
            if seenNum:
                addy += ' ' +  row[i].value #adding a space so the address will be better formatted
        customer.ship2addy = addy
        
            #getting the bill2addy
        addy = ''
        seenNum = False
        for i in bill2addy:
            if row[i].value == None: continue #skipping cells that have None value
            if ord(row[i].value[0]) <= ord('9') or ord(row[i].value[0]) >= ord('0'): #this cell starts with a number, this is checked so that all the cases where the customer name is first-second cell get skipped
                seenNum = True
            if seenNum:
                addy += ' ' +  row[i].value
        customer.bill2addy = addy
        
        #if the ship2addy is a real address, we want to save it as the useAddy, else try bill2addy.  If neither just have useAddy be an empty string
        lat, lng = getCoordinates(customer.ship2addy)
        if lat != None: #only have to check one because will always get both if the address is valid
            customer.useAddy = customer.ship2addy
            customer.lng = lng
            customer.lat = lat
            customers.append(customer)
            continue#process next row
        else: #the ship2addy didn't give a result
            lat, lng = getCoordinates(customer.bill2addy)
            if lat != None:
                customer.useAddy = customer.bill2addy
                customer.lng = lng
                customer.lat = lat
                customers.append(customer)
                continue #process next row
            else: #both addresses were invalid
                customer.useAddy = ''
                customer.lng = None
                customer.lat = None
                customers.append(customer)
    workbook.close() #done reading from this workbook    
    
    #Now to write back to output file
    newbook = Workbook()
    newsheet = newbook.active
    
    #adding headder row to newsheet
    for row in newsheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(headings)):
        for i in range(len(headings)):
            row[i].value = headings[i]

    #adding data for each potty in the list to the newbook
    i = 0
    for row in newsheet.iter_rows(min_row=2, max_row=len(customers)+1, min_col=1, max_col=len(headings)):
        #manually assign each colum in row, unless later i decide to just keep everything in a list (will need to decide which is better, easy acces with name of varable or memorizing the position of the variable in a list)
        customer = customers[i]
        row[0].value = customer.name
        row[1].value = customer.useAddy
        row[2].value = 'Monday'
        # row[3].value = customer.notes    #I am not extracting note information from the original file
        row[4].value = customer.lat
        row[5].value = customer.lng
        good = customer.useAddy != '' #saving whether the address was good or not
        row[6].value = good
        i += 1
    

    #if lat and lng were added to some potties, we need to save the excel file
    newbook.save(filename=firstOutput)
    newbook.close() #close workbook
    return





#testing
FirstLoad()