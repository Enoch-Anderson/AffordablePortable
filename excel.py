from Classes import Potty
from openpyxl import Workbook, load_workbook
from Mappings import name, address, cleaningDay, longitude, latitude, notes, openfile, outputfiles, weekdayToIntDictionary, intToWeekday, headers
from GoogleStuff import getCoordinates

#Tutorial on how to use openpyxl: https://realpython.com/openpyxl-excel-spreadsheets-python/

def weekdayToInt(weekday: str):
    return weekdayToIntDictionary[weekday]

def loadExcel():
    #Load workbook, get ready to access data
    workbook = load_workbook(filename=openfile) #didn't set readonly to true becasue will add lat and lng for potties without them already
    sheet = workbook.active
    #grab headers for use latter
    for cell in sheet[1]:
        headers.append(cell.value)

    #access data and store data in a list of Potty classes
    potties = []
    i = 0
    for row in sheet.iter_rows(min_row=2):
        if row[name].value == None: break #This breaks out if the happen to have other stuff in excel file on accident
        #if the lat, lng are not already in the excell file, will use the gmpas api to get them and save them
        if row[longitude].value == None or row[latitude].value == None:
            lng, lat = getCoordinates(row[address].value)#get values
            row[longitude].value = lng #set the values
            row[latitude].value = lat #set the values
        potty = Potty(number= i, #number is the row number that it appears in the excel file, no loger an input in the excel file 
                    name=row[name].value, 
                    address=row[address].value,
                    cleanDayInt= weekdayToIntDictionary[ row[cleaningDay].value ], #Get the cleanind day string from excel and convert it to the int to be used later
                    longitude=row[longitude].value, 
                    latitude=row[latitude].value, 
                    notes=row[notes].value,
                    badAddress=False)
        if potty.longitude == None or potty.latitude == None:
            potty.badAddress = True 
        potties.append(potty) #Make it so that the badAddresses get appended to this, and there is a checker for good addresses in the folium map later
        i += 1

    #if lat and lng were added to some potties, we need to save the excel file
    workbook.save(filename=openfile)
    workbook.close() #close workbook
    return potties


def SaveAll(potties, weekdayArrays):
    #save potties to main file first
    saveOne(potties, openfile)
    
    #save each week day to week day file
    for i, lst in enumerate(weekdayArrays):
        saveOne(lst, outputfiles[i])
    
    return

def saveOne(lst, savefile):
    #saving potties that need to be cleaned today to a new excell spreadsheet
    newbook = Workbook()
    newsheet = newbook.active

    #adding headder row to newsheet
    for row in newsheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for i in range(len(headers)):
            row[i].value = headers[i]

    #adding data for each potty in the list to the newbook
    i = 0
    for row in newsheet.iter_rows(min_row=2, max_row=len(lst)+1, min_col=1, max_col=len(headers)):
        #manually assign each colum in row, unless later i decide to just keep everything in a list (will need to decide which is better, easy acces with name of varable or memorizing the position of the variable in a list)
        potty = lst[i]
        row[0].value = potty.name
        row[1].value = potty.address
        row[2].value = intToWeekday[potty.cleanDayInt]
        row[3].value = potty.notes
        row[4].value = potty.latitude
        row[5].value = potty.longitude
        row[6].value = (not potty.badAddress)
        i += 1

    #save newbook to commit changes to it
    newbook.save(filename=savefile)
    newbook.close()
    return


# #This is how you clustered stuff for later use
# #cluster the potties for display
# clustered = clusterPotties(potties, numerOfClusters)
# #create and display map
# CreateAndDisplayMap(clustered)
